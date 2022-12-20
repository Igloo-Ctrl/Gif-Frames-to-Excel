from openpyxl import Workbook, load_workbook
from PIL import Image
import os

gif_path = "/Users/ryanjones/Downloads/classic-dancing-banana.gif"

try:
    os.mkdir("GIF Frames")
except FileExistsError:
    pass

image = Image.open(gif_path)
number_of_frames = image.n_frames

for i in range(number_of_frames):
    image.seek(i)
    image.save(f"GIF Frames/frame_{i + 1}.gif")
    print(f"Frame {i + 1} saved!")

# desired Excel filename
excel_filename = "banana"

# path of the folder containing the images
folder_name = "GIF Frames"

amount_of_images = len(os.listdir(folder_name))

wb = Workbook()
wb.save(filename=f"{excel_filename}.xlsx")
for i in range(1, amount_of_images + 1):
    wb.create_sheet(title=f"Frame {i}")
del wb["Sheet"]
wb.save(filename=f"{excel_filename}.xlsx")

# create frames folder
try:
    os.mkdir("RGB Values")
    print('"RGB Values" folder created.')
except FileExistsError:
    pass

print("Creating text files containing RGB values.")

folder_contents = os.listdir(folder_name)
for image in enumerate(folder_contents):

    # open the image, convert to rgb and then load it
    scene_image = Image.open(f"{folder_name}/{image[1]}")
    rgb_image = scene_image.convert("RGB")

    # image width and height
    image_width, image_height = scene_image.size[0], scene_image.size[1]
    # print(image_width, image_height)

    # string to place all the data
    data_string = ""

    # nested for loop grab every pixel from the image, grab its specific rgb value and slaps it into a document
    for i in range(1, image_height):
        for j in range(1, image_width):
            data_string += f"" \
                           f"{rgb_image.getpixel((j, i))[0]}, {rgb_image.getpixel((j, i))[1]}, " \
                           f"{rgb_image.getpixel((j, i))[2]}\n"

    # writes to document
    with open(f"RGB Values/RGB_{image[0] + 1}.txt", "w") as f:
        f.write(data_string)
        print(f"RGB_{image[0] + 1}.txt complete!")

    rgb_image.close()

document_number = 0
workbook = load_workbook(filename=f"{excel_filename}.xlsx")

print("Passing RGB values into the Excel document.")

for filename in os.listdir("RGB Values"):
    row, column = 1, 1
    workbook.active = document_number
    sheet = workbook.active
    with open(f"RGB Values/{filename}", "r") as f:
        data = f.readlines()
        for i in data:
            if column > image_width - 1:
                row += 1
                column = 1
            sheet.cell(column=column, row=row, value=i)
            column += 1
        print(f"Sheet {document_number + 1} complete!")
        document_number += 1

print("Saving all the changes... ")
workbook.save(filename=f"{excel_filename}.xlsx")
print("Complete!")
