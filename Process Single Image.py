import time
import os
from os.path import exists
import uuid
from PIL import Image
from openpyxl import Workbook, load_workbook

# enter the file path to your image here, for example, /Users/johndoe/Downloads/cat.png
desired_filepath = "Cat-Photog-Feat-256x256.jpg"

# time taken
start = time.process_time()


class Variables:

    image_width = 0
    image_height = 0

    image_filepath = ""
    image_filename = ""

    excel_file_path = ""


def check_for_folder():
    """
    check for the images folder in the current directory and if it isn't found, creates it
    """
    print(f'Checking "{os.getcwd()}" for "images" folder.')
    if exists("images"):
        print('"images" folder found.')
    else:
        os.mkdir("images")
        print('"images" folder not found. Creating.')
    validate_filepath()


def validate_filepath():
    """
    checks to see if the filepath is empty, if so exits the program, otherwise proceeds with the process
    catches an IOError if a valid image doesn't exist at the filepath or if anything below is poorly configured
    """
    if desired_filepath == "":
        exit('Filepath empty, please change the filepath at the top of this script to a valid location.')
    else:
        try:
            image = Image.open(desired_filepath)
            image.close()
            print(f'Valid image found at "{desired_filepath}". Proceeding.')
            create_and_move(desired_filepath)
            process_image(desired_filepath)
        except IOError:
            print("Invalid file type or file not found.")


def create_and_move(filepath):
    """
    stores information in the Variables class, creates a unique directory for the process and makes a copy of
    the original image there
    """
    # directory
    Variables.image_filename = os.path.basename(filepath)
    generated_filename = str(uuid.uuid4())
    Variables.image_filepath = f"images/{generated_filename}"
    os.mkdir(Variables.image_filepath)

    # saving a copy of the image
    image = Image.open(filepath)
    image.save(f"images/{generated_filename}/{Variables.image_filename}")
    image.close()
    print("Creating a folder for this process and copying your image there.")


def process_image(filepath):
    """
    iterates through the pixels of the image, grabbing their RGB values and adding them to a big string
    finally, this string is saved in a text document
    """
    open_image = Image.open(filepath)
    Variables.image_width, Variables.image_height = open_image.size[0], open_image.size[1]
    # rgb_image = open_image.convert("rgb")

    data_string = ""
    # i = height, j = width (??)
    for i in range(Variables.image_height):
        for j in range(Variables.image_width):
            data_string += f"{open_image.getpixel((j, i))[0]}, {open_image.getpixel((j, i))[1]}, " \
                           f"{open_image.getpixel((j, i))[2]}\n"

    with open(f"{Variables.image_filepath}/rgb.txt", "w") as f:
        f.write(data_string)
        print(f"RGB text file created.")
    create_work_book()


def create_work_book():
    """
    creates a workbook, deletes the first sheet (typically named Sheet) and creates a new one based off
    the image name
    """
    wb = Workbook()
    wb.create_sheet(title=f"{Variables.image_filename}")
    del wb["Sheet"]
    Variables.excel_file_path = f"{Variables.image_filepath}/excel.xlsx"
    wb.save(Variables.excel_file_path)
    print("Excel Workbook created.")
    input_rbg_into_excel()


def input_rbg_into_excel():
    """
    reads the text file and iterates through it, inputting the values into the Excel sheet
    """
    workbook = load_workbook(Variables.excel_file_path)
    row, column = 1, 1
    sheet = workbook.active
    with open(f"{Variables.image_filepath}/rgb.txt", "r") as f:
        data = f.readlines()
        for i in data:
            if column > Variables.image_width:
                row += 1
                column = 1
            sheet.cell(column=column, row=row, value=i)
            column += 1
    workbook.save(Variables.excel_file_path)
    print("RBG values inputted successfully.")
    write_vba_macros()


def write_vba_macros():
    """
    writes VBA macros specific to the image used, the first one concerns changing the cell size which is more
    general, the second is used for colouring the sheet
    """
    comment = f"'These macros were created for the image: {Variables.image_filename}. Run the first to resize " \
              f"the cells, run the second to colour them in.\n"
    macro_one = """Sub ChangeColumnWidth()
'Changes the column width of the range of cells to 2

    Columns("A:ZZ").ColumnWidth = 2

End Sub
    """
    macro_two = f"""
Sub ConvertCellValuesToBackground()
'Grabs the cell value, splits it at the comma and then using a nested for loop, feeds the values into an RGB value
'Used in the creation of the frames

For i = 1 To {Variables.image_width}
    For j = 1 To {Variables.image_height}
        
        myArray = Split(Cells(j, i).Value, ", ")
        Cells(j, i).Interior.Color = RGB(myArray(0), myArray(1), myArray(2))
        
    Next j
Next i

End Sub
    """
    compiled_text = f"{comment}{macro_one}{macro_two}"
    with open(f"{Variables.image_filepath}/vba_macros.txt", "w") as f:
        f.write(compiled_text)
    print("Finished writing VBA macros.")
    print(F"All done! Time taken: {time.process_time() - start} seconds.")


def main():
    check_for_folder()


if __name__ == '__main__':
    main()
