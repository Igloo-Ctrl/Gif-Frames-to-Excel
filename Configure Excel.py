"""the use of this script depends on a gif still that is equal in both width and height"""

import os
from openpyxl import load_workbook, Workbook

# openpyxl
workbook = load_workbook(filename='hello_world.xlsx')

# lists the files within the frames directory
list_of_files = sorted(os.listdir("frames"), key=len)


class Variables:
    # change this number depending on the width of your image
    width = 48


class documentNumber:
    document = 0


def populate_worksheet_with_rgb_values(name_of_file):
    """
    iterates through the document conjured up by refactor.py, places the line's values in a cell and increment the
    column count
    when the image's width is reached, the row is changed and the column is reset
    """
    row, column = 1, 1
    workbook.active = documentNumber.document
    sheet = workbook.active
    with open(f"frames/{name_of_file}", "r") as f:
        data = f.readlines()
        for i in data:
            if column > Variables.width - 1:
                row += 1
                column = 1
            sheet.cell(column=column, row=row, value=i)
            column += 1
        documentNumber.document += 1


def main():
    for filename in list_of_files:
        populate_worksheet_with_rgb_values(filename)
        print(f"{filename} complete! Moving on.")
    workbook.save(filename='hello_world.xlsx')


# sheet creation and manipulation
def create_work_book():
    wb = Workbook()
    wb.save(filename='hello_world.xlsx')


def create_sheets():
    for i in range(1, 58):
        workbook.create_sheet(title=f"Frame {i}")


if __name__ == '__main__':
    main()
